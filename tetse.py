import openpyxl
from openpyxl import cell

def gerarColunas():
    import string
    a = string.ascii_uppercase
    lst = []
    for i in a:
        lst.append(i)
    lst2 = []
    for j in range(5):
        for i in a:
            lst2.append(lst[j] + i)
    lst += lst2
    return lst

wb = openpyxl.load_workbook("Resultado.xlsx", read_only=False, data_only=True)
ws = wb.active
colunas = gerarColunas()

for i in colunas:
    if ws[i + "1"].value == "Solicitação":
        coluna_solicitacao = i

    elif ws[i + "1"].value == "Situação":
        coluna_situacao = i

    elif ws[i + "1"].value == "Localização":
        coluna_localizacao = i

    elif ws[i + "1"].value == "Início":
        coluna_inicio = i

    elif ws[i + "1"].value == "Fim":
        coluna_fim = i

    elif ws[i + "1"].value == "SLACONSUMIDOEXECUCAOTEC - solicitacao_infraestrutura":
        coluna_slaexe = i

    elif ws[i + "1"].value == "SLACONSUMIDOACOMPANHAMENTO - solicitacao_infraestrutura":
        coluna_slaacom = i

    elif ws[i + "1"].value == "DESCRICAOITEM - solicitacao_infraestrutura":
        coluna_item = i

    elif ws[i + "1"].value == "TOTALHOURSSLA - solicitacao_infraestrutura":
        coluna_sla = i

    elif ws[i + "1"].value == "CODIGOTECNICO - solicitacao_infraestrutura":
        coluna_tec = i



wb.close()

