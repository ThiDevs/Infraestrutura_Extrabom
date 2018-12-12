import openpyxl
import datetime

def PadronizaDate(date):
    date = date.split("-")[0] + "/" + date.split("-")[1] + "/" + date.split("-")[2].split("T")[0] + " " + date.split("T")[1]
    return date

def PadronizaDate2(date):
    date = date.split("/")[2].split(" ")[0] + "/" +date.split("/")[1] + "/" + date.split("/")[0] + " " + date.split(" ")[1]
    return date


def calcHours(s,t):
    from datetime import datetime
    f = '%Y/%m/%d %H:%M:%S'
    dif = (datetime.strptime(t, f) - datetime.strptime(s, f)).total_seconds()
    # print(dif/60/60)
    return dif/60/60

def main():

    wb = openpyxl.load_workbook("Resultado.xlsx", read_only=False)
    ws = wb.active

    z = 0
    Dentro_do_Prazo = 0

    for i in range(2,len(ws['B']),1):#len(ws['B'])):

        finalizada = str(ws['C' + str(i)].value).strip()
        solicitacao = str(ws['A' + str(i)].value).strip()

        if finalizada == "Finalizada":

            DataEntrada_classificacao = str(ws['AH' + str(i)].value).strip()          #DATAENTRADACLASSIFICACAO - solicitacao_infraestrutura

            Data_Entrada_AprovarConclusao = PadronizaDate2(str(ws['DS' + str(i)].value).strip())  # Atividade - solicitacao_infraestrutura - Aprovar conclusão - Conclusão

            Item_Sla = str(ws['AE' + str(i)].value).strip().split(":")[0] #TOTALHOURSSLA - solicitacao_infraestrutura

            if DataEntrada_classificacao != "" and DataEntrada_classificacao != None and DataEntrada_classificacao != " " and DataEntrada_classificacao != "None":

                dataEntrada = PadronizaDate(DataEntrada_classificacao)

                dif = calcHours(dataEntrada + ":00", Data_Entrada_AprovarConclusao)

                if dif >= int(Item_Sla):
                    pass
                else:
                    Dentro_do_Prazo += 1

            else:

                """ Com orgão fiscalizador ativo """
                DataSaida_Inicio = str(ws['Q'+str(i)].value).strip() #Atividade - solicitacao_infraestrutura - Início - Conclusão

                DataSaida_Exec = str(ws['AL' + str(i)].value).strip()  # DATASAIDAEXECUCAOTEC - solicitacao_infraestrutura

                if DataSaida_Exec != "" and DataSaida_Exec != None and DataSaida_Exec != " " and DataSaida_Exec != "None":

                    dataSaida_Exec = PadronizaDate(DataSaida_Exec) + ":00"

                else:
                    DataSaida_Exec = str(ws['AO' + str(i)].value).strip() # DATASAIDAACOMPANHAMENTO - solicitacao_infraestrutura

                    dataSaida_Exec = PadronizaDate(DataSaida_Exec) + ":00"

                dataSaida_Inicio = PadronizaDate2(DataSaida_Inicio)
                dif = calcHours(dataSaida_Inicio, dataSaida_Exec)

                if dif >= int(Item_Sla):
                    pass
                else:
                    Dentro_do_Prazo += 1
            z += 1
        print("Dentro do prazo", Dentro_do_Prazo)




main()
#2018-12-04T16:06


#AL OU AO

















