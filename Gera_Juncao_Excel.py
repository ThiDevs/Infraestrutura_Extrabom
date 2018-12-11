from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors


def Juncao():
    wb = load_workbook(filename='Relatorio/Relatorio_InternoVsExterno.xlsx', read_only=False)
    ws = wb['Dados']
    dic_newExcel = {}
    percorre = 200
    for i in range(2, percorre, 1):

        solicitacoes = str(ws['A' + str(i)].value)
        tecnico = ws['D' + str(i)].value
        DentroPrazo = str(ws['J' + str(i)].value).strip()
        situacao = str(ws['B' + str(i)].value).strip()
        if situacao == "Fechado":
            try:
                lst = dic_newExcel.get(tecnico)[0]
                lst.append(solicitacoes)
                prazo = dic_newExcel.get(tecnico)[1]

            except Exception:
                lst = [solicitacoes]
                prazo = 0

            if DentroPrazo == "Sim":
                prazo += 1
            dic_newExcel.update({tecnico: (lst, prazo)})

    print(dic_newExcel)

    ws_Juntos = wb.create_sheet(title="Junção")

    ft = Font(color=colors.WHITE)

    blackFill = PatternFill(start_color='010204',
                            end_color='010204',
                            fill_type='solid')

    lista_names = ['Técnico', 'Total de chamados', 'Dentro do prazo', 'Porcetagem', 'Fora do prazo']
    ws_Juntos.auto_filter.ref = 'A1:E9'
    z = 1
    for name in lista_names:
        ws_Juntos.cell(row=1, column=z).value = name
        ws_Juntos.cell(row=1, column=z).font = ft
        ws_Juntos.cell(row=1, column=z).fill = blackFill
        z += 1

    names = ['Paulo Tavares', 'Bruno Soares', 'Romildo Carvalho', 'Aires Mendonça', 'Marcos Aurélio', 'Valcleide Silva',
             'Daniel Ferreira', 'Julio Pulcher', 'Andre Melo', 'Sandro Geraldino', 'Claudio']

    names_Eletrica = ['marcelo.silva', 'carlos.simoes', 'ubirajara.silva']

    names_Mecanica = ['gean.vieira']

    dic_Civil = {}
    dic_Eletrica = {}
    dic_Mecanica = {}

    for key in dic_newExcel.keys():
        for name in names:
            if name in key:
                dic_Civil.update({key: len(dic_newExcel.get(key)[0])})

        for name in names_Eletrica:
            if name in key:
                dic_Eletrica.update({key: len(dic_newExcel.get(key)[0])})

        for name in names_Mecanica:
            if name in key:
                dic_Mecanica.update({key: len(dic_newExcel.get(key)[0])})

        len_soli = len(dic_newExcel.get(key)[0])
        ws_Juntos.append(
            [key, len_soli, dic_newExcel.get(key)[1], str(round((dic_newExcel.get(key)[1] * 100) / len_soli)) + " %",
             len_soli - dic_newExcel.get(key)[1]])

    total = 0
    linha = 7

    ft = Font(color="c65911")

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         bottom=Side(style='thin'), top=Side(style='thin'))

    ws = wb['Informações']

    ws.cell(row=6, column=1).value = "Técnicos Civis"
    ws.cell(row=6, column=1).border = thin_border
    ws.cell(row=6, column=1).font = ft

    ws.cell(row=6, column=2).value = "Solicitações"
    ws.cell(row=6, column=2).border = thin_border
    ws.cell(row=6, column=2).font = ft

    ordena = list(reversed(sorted(dic_Civil.values())))
    for i in range(len(dic_Civil.keys())):
        for names in dic_Civil.keys():
            if dic_Civil[names] == ordena[i]:
                cell = ws.cell(row=linha, column=1)
                cell.value = names
                cell.font = ft
                cell.border = thin_border

                cell2 = ws.cell(row=linha, column=2)
                cell2.value = ordena[i]
                cell2.font = ft
                cell2.border = thin_border

                total += int(ordena[i])
                linha += 1
                break
        del dic_Civil[names]
    ws.cell(row=linha, column=1).value = "Total"
    ws.cell(row=linha, column=1).font = ft
    ws.cell(row=linha, column=1).border = thin_border

    ws.cell(row=linha, column=2).value = total
    ws.cell(row=linha, column=2).font = ft
    ws.cell(row=linha, column=2).border = thin_border

########################################################################################################################

    ws.cell(row=6, column=4).value = "Técnicos Elétrica"
    ws.cell(row=6, column=4).border = thin_border
    ws.cell(row=6, column=4).font = ft

    ws.cell(row=6, column=5).value = "Solicitações"
    ws.cell(row=6, column=5).border = thin_border
    ws.cell(row=6, column=5).font = ft

    total = 0
    linha = 7

    ordena = list(reversed(sorted(dic_Eletrica.values())))
    for i in range(len(dic_Eletrica.keys())):
        for names in dic_Eletrica.keys():
            if dic_Eletrica[names] == ordena[i]:
                cell = ws.cell(row=linha, column=4)
                cell.value = names
                cell.font = ft
                cell.border = thin_border

                cell2 = ws.cell(row=linha, column=5)
                cell2.value = ordena[i]
                cell2.font = ft
                cell2.border = thin_border

                total += int(ordena[i])
                linha += 1
                break
        del dic_Eletrica[names]

    ws.cell(row=linha, column=4).value = "Total"
    ws.cell(row=linha, column=4).font = ft
    ws.cell(row=linha, column=4).border = thin_border

    ws.cell(row=linha, column=5).value = total
    ws.cell(row=linha, column=5).font = ft
    ws.cell(row=linha, column=5).border = thin_border


########################################################################################################################

    ws.cell(row=6, column=7).value = "Técnicos Mecânica"
    ws.cell(row=6, column=7).border = thin_border
    ws.cell(row=6, column=7).font = ft

    ws.cell(row=6, column=8).value = "Solicitações"
    ws.cell(row=6, column=8).border = thin_border
    ws.cell(row=6, column=8).font = ft

    total = 0
    linha = 7

    ordena = list(reversed(sorted(dic_Mecanica.values())))
    for i in range(len(dic_Mecanica.keys())):
        for names in dic_Mecanica.keys():
            if dic_Mecanica[names] == ordena[i]:
                cell = ws.cell(row=linha, column=7)
                cell.value = names
                cell.font = ft
                cell.border = thin_border

                cell2 = ws.cell(row=linha, column=8)
                cell2.value = ordena[i]
                cell2.font = ft
                cell2.border = thin_border

                total += int(ordena[i])
                linha += 1
                break
        del dic_Mecanica[names]

    ws.cell(row=linha, column=7).value = "Total"
    ws.cell(row=linha, column=7).font = ft
    ws.cell(row=linha, column=7).border = thin_border

    ws.cell(row=linha, column=8).value = total
    ws.cell(row=linha, column=8).font = ft
    ws.cell(row=linha, column=8).border = thin_border


    wb.save('Relatorio/Relatorio_InternoVsExterno.xlsx')
    wb.close()


Juncao()
