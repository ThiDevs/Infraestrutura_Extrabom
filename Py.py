from tkinter import *
from tkinter import filedialog
from openpyxl import load_workbook


class Application:
    def __init__(self, master=None):
        self.widget1 = Frame(master)
        self.widget1.pack()
        self.msg = Label(self.widget1, text="Primeiro widget")
        self.msg["font"] = ("Calibri", "9", "italic")
        self.msg.pack()
        self.sair = Button(self.widget1)
        self.sair["text"] = "Exportar Excel"
        self.sair["font"] = ("Calibri", "9")
        self.sair["width"] = 10
        self.sair.bind("<Button-1>", self.mudarTexto)
        self.sair.pack()

    def mudarTexto(self, event):
        import threading
        threading.Thread(target=file, args=(self,)).start()


def file(self):
    self.filename = filedialog.askopenfilename(initialdir="/", title="Select file",
                                               filetypes=(("Excel", "*.xlsx"), ("all files", "*.*")))
    print(self.filename)
    import threading
    threading.Thread(target=ExportaExcel, args=(self, self.filename,)).start()


def ExportaExcel(self, filename):
    wb = load_workbook(filename=filename, read_only=True)
    ws = wb['Resultado da consulta de solici']
    print()

    arq = open("Tecnicos.txt", 'rt')
    lines = arq.readlines()
    dic = {}
    for line in lines:
        dic.update({line.split(";")[0]: line.split(";")[1].strip()})
    arq.close()

    from openpyxl import Workbook
    wb_new = Workbook(write_only=True)

    ws_new = wb_new.create_sheet()

    ws_new.append(
        ['Solicitações', 'Situação', 'Localização', 'Técnico', 'Inicio', 'Fim', 'Item', 'SLA', 'SLA Consumido',
         "Dentro do prazo?"])
    dic_newExcel = {}

    for i in range(2, 50, 1):

        solicitacoes = str(int(ws['A' + str(i)].value))
        localizacao = str(ws['F' + str(i)].value).strip()
        if localizacao == "Classificação" or localizacao == "Acompanhamento da Execução" or localizacao == "Acompanharmento da Execução" or localizacao == "Execução da Solicitação":
            situacao = "Em aberto"
        elif localizacao == "Aprovar":
            situacao = "Aprovação"
        else:
            situacao = "Fechado"

        try:
            tecnico = dic[ws['AS' + str(i)].value]
        except Exception:
            tecnico = ws['AS' + str(i)].value

        item = ws['BU' + str(i)].value
        SLA = ws['AA' + str(i)].value
        inicio = ws['H' + str(i)].value
        fim = ws['I' + str(i)].value

        SLA_Consumido = ""
        SLA_Consumido2 = ""

        if ws['AG' + str(i)].value != '':
            SLA_Consumido = ws['AF' + str(i)].value.split(":")[0] + ":00"  # SLA EXECUÇÃO#

        if ws['AJ' + str(i)].value != '':
            SLA_Consumido2 = ws['AI' + str(i)].value.split(":")[0] + ":00"  # SLA ACOMPANHAMENTO#

        if SLA_Consumido != "" and SLA_Consumido2 != "":
            if int(SLA_Consumido.split(":")[0]) < int(SLA_Consumido2.split(":")[0]):
                SLA_Consumido = SLA_Consumido2
        else:
            SLA_Consumido = "00:00"

        try:
            lst = dic_newExcel.get(tecnico)[0]
            lst.append(solicitacoes)

            prazo = dic_newExcel.get(tecnico)[1]

        except Exception:
            lst = []
            lst.append(solicitacoes)

            prazo = 0

        if int(SLA_Consumido.split(":")[0]) < int(SLA.split(":")[0]) and SLA_Consumido != "00:00":
            ws_new.append(
                [solicitacoes, situacao, localizacao, tecnico, inicio, fim, item, SLA, SLA_Consumido, "Sim"])
            prazo += 1
        elif int(SLA_Consumido.split(":")[0]) > int(SLA.split(":")[0]):
            ws_new.append(
                [solicitacoes, situacao, localizacao, tecnico, inicio, fim, item, SLA, SLA_Consumido, "Não"])
        else:
            ws_new.append(
                [solicitacoes, situacao, localizacao, tecnico, inicio, fim, item, SLA, SLA_Consumido, ""])

        dic_newExcel.update({tecnico: (lst, prazo)})
        self.msg['text'] = str(int((i * 100) / 1000)) + "%"
        print(int((i * 100) / 1000), "%")

    ws_new.close()
    wb_new.save('New Excel Infra.xlsx')
    wb_new.close()


def main():
    root = Tk()
    Application(root)
    root.mainloop()


main()
