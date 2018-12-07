import datetime
from openpyxl import load_workbook

dic = {}


def main():

    wb = load_workbook(filename='Relatorio/Relatorio_InternoVsExterno.xlsx', read_only=False)
    ws = wb['Dados']
    dic_newExcel = {}
    percorre = 200
    for i in range(2, percorre, 1):

        solicitacoes = str(ws['A' + str(i)].value)
        tecnico = ws['D' + str(i)].value
        DentroPrazo = str(ws['J' + str(i)].value).strip()
        situacao = str(ws['B' + str(i)].value).strip()
        if situacao == "Fechado":
            try:
                lst = dic_newExcel.get(tecnico)[0]
                lst.append(solicitacoes)
                prazo = dic_newExcel.get(tecnico)[1]

            except Exception:
                lst = [solicitacoes]
                prazo = 0

            if DentroPrazo == "Sim":
                prazo += 1
            dic_newExcel.update({tecnico: (lst, prazo)})

    print(dic_newExcel)

    ws_Juntos = wb.create_sheet(title="Junção")
    ws_Juntos.append(['Técnico', 'Total de chamados', 'Dentro do prazo', 'Porcetagem', 'Fora do prazo'])
    for key in dic_newExcel.keys():
        print(key)

        len_soli = len(dic_newExcel.get(key)[0])
        ws_Juntos.append([key, len_soli, dic_newExcel.get(key)[1], str(round((dic_newExcel.get(key)[1]*100) / len_soli)) + " %", len_soli - dic_newExcel.get(key)[1]])

    wb.save('new_big_file23.xlsx')
    wb.close()


main()
