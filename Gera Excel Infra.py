import datetime
from openpyxl import load_workbook


def main():
    wb = load_workbook(filename='Resultado da consulta de solicitações.xlsx', read_only=True)
    ws = wb['Resultado da consulta de solici']

    arq = open("Tecnicos.txt", 'rt')
    lines = arq.readlines()
    dic = {}
    for line in lines:
        dic.update({line.split(";")[0]: line.split(";")[1].strip()})
    arq.close()

    from openpyxl import Workbook
    wb_new = Workbook(write_only=True)

    ws_new = wb_new.create_sheet()

    ws_new.append(['Solicitações', 'Técnico', 'Inicio', 'Fim', 'Prazo', 'Item', 'SLA', "Dentro do prazo?"])
    dic_newExcel = {}

    for i in range(496):
        try:
            if ws['BC' + str(i)].value.isnumeric():
                solicitacoes = str(int(ws['A' + str(i)].value))

                tecnico = dic[ws['BC' + str(i)].value]

                horas = calcdays(ws['H' + str(i)].value, ws['I' + str(i)].value)
                item = ws['CE' + str(i)].value
                SLA = ws['AK' + str(i)].value
                inicio = ws['H' + str(i)].value
                fim = ws['I' + str(i)].value
                try:
                    lst = dic_newExcel.get(tecnico)[0]
                    lst.append(solicitacoes)

                    prazo = dic_newExcel.get(tecnico)[1]

                except Exception:
                    lst = []
                    lst.append(solicitacoes)

                    prazo = 0


                if int(horas) < int(SLA.split(":")[0]):
                    ws_new.append([solicitacoes, tecnico, inicio, fim, horas, item, SLA, "Sim"])
                    prazo += 1


                else:
                    ws_new.append([solicitacoes, tecnico, inicio, fim, horas, item, SLA, "Não"])

                dic_newExcel.update({tecnico: (lst, prazo)})


        except Exception:
            pass
        print(int((i*100)/496), "%")

    ws_new.close()


    ws_Juntos = wb_new.create_sheet(title="Junção")
    ws_Juntos.append(['Técnico', 'Total de chamados', 'Dentro do prazo', 'Fora do prazo'])
    for key in dic_newExcel.keys():
        print(key)

        len_soli = len(dic_newExcel.get(key)[0])
        ws_Juntos.append([key, len_soli, dic_newExcel.get(key)[1], len_soli - dic_newExcel.get(key)[1]])

    wb_new.save('new_big_file23.xlsx')
    wb_new.close()

def calcdays(date1, date2):
    day = date1.split('/')[0]
    month = date1.split('/')[1]
    year = date1.split('/')[2]
    data1 = datetime.date(day=int(day), month=int(month), year=int(year))
    data2 = datetime.date(day=int(date2.split('/')[0]), month=int(date2.split('/')[1]), year=int(date2.split('/')[2]))
    diferenca = data2 - data1
    return str(diferenca.days * 24)


main()
