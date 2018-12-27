import datetime
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors

dic = {}


def PadronizaDate(date):
    date = date.split("-")[0] + "/" + date.split("-")[1] + "/" + date.split("-")[2].split("T")[0] + " " + \
           date.split("T")[1]
    return date


def PadronizaDate2(date):
    try:
        date = date.split("/")[2].split(" ")[0] + "/" + date.split("/")[1] + "/" + date.split("/")[0] + " " + \
               date.split(" ")[1]
    except Exception:
        pass
    return date


def calcHours(s, t):
    from datetime import datetime
    f = '%Y/%m/%d %H:%M:%S'
    dif = (datetime.strptime(t, f) - datetime.strptime(s, f)).total_seconds()
    # print(dif/60/60)
    return dif / 60 / 60


def Juncao():
    wb = load_workbook(filename='Relatorio/Relatorio_InternoVsExterno.xlsx', read_only=False)
    ws = wb['Dados']
    dic_newExcel = {}
    percorre = 200
    for i in range(2, percorre, 1):

        solicitacoes = str(ws['A' + str(i)].value)
        tecnico = ws['D' + str(i)].value
        DentroPrazo = str(ws['J' + str(i)].value).strip()
        situacao = str(ws['B' + str(i)].value).strip()
        if situacao == "Fechado":
            try:
                lst = dic_newExcel.get(tecnico)[0]
                lst.append(solicitacoes)
                prazo = dic_newExcel.get(tecnico)[1]

            except Exception:
                lst = [solicitacoes]
                prazo = 0

            if DentroPrazo == "Sim":
                prazo += 1
            dic_newExcel.update({tecnico: (lst, prazo)})

    ws_Juntos = wb.create_sheet(title="Junção")

    ft = Font(color=colors.WHITE)

    blackFill = PatternFill(start_color='010204',
                            end_color='010204',
                            fill_type='solid')

    lista_names = ['Técnico', 'Total de chamados', 'Dentro do prazo', 'Porcetagem', 'Fora do prazo']
    ws_Juntos.auto_filter.ref = 'A1:E9'
    z = 1
    for name in lista_names:
        ws_Juntos.cell(row=1, column=z).value = name
        ws_Juntos.cell(row=1, column=z).font = ft
        ws_Juntos.cell(row=1, column=z).fill = blackFill
        z += 1

    names = ['Paulo Tavares', 'Bruno Soares', 'Romildo Carvalho', 'Aires Mendonça', 'Marcos Aurélio', 'Valcleide Silva',
             'Daniel Ferreira', 'Julio Pulcher', 'Andre Melo', 'Sandro Geraldino', 'Claudio']

    names_Eletrica = ['marcelo.silva', 'carlos.simoes', 'ubirajara.silva']

    names_Mecanica = ['gean.vieira']

    dic_Civil = {}
    dic_Eletrica = {}
    dic_Mecanica = {}

    for key in dic_newExcel.keys():
        for name in names:
            if name in key:
                dic_Civil.update({key: len(dic_newExcel.get(key)[0])})

        for name in names_Eletrica:
            if name in key:
                dic_Eletrica.update({key: len(dic_newExcel.get(key)[0])})

        for name in names_Mecanica:
            if name in key:
                dic_Mecanica.update({key: len(dic_newExcel.get(key)[0])})

        len_soli = len(dic_newExcel.get(key)[0])
        ws_Juntos.append(
            [key, len_soli, dic_newExcel.get(key)[1], str(round((dic_newExcel.get(key)[1] * 100) / len_soli)) + " %",
             len_soli - dic_newExcel.get(key)[1]])

    total = 0
    linha = 7

    ft = Font(color="c65911")

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         bottom=Side(style='thin'), top=Side(style='thin'))


    ws = wb['Informações']

    ws.cell(row=6, column=1).value = "Técnicos Civis"
    ws.cell(row=6, column=1).border = thin_border
    ws.cell(row=6, column=1).font = ft

    ws.cell(row=6, column=2).value = "Solicitações"
    ws.cell(row=6, column=2).border = thin_border
    ws.cell(row=6, column=2).font = ft

    ordena = list(reversed(sorted(dic_Civil.values())))
    for i in range(len(dic_Civil.keys())):
        for names in dic_Civil.keys():
            if dic_Civil[names] == ordena[i]:
                cell = ws.cell(row=linha, column=1)
                cell.value = names
                cell.font = ft
                cell.border = thin_border

                cell2 = ws.cell(row=linha, column=2)
                cell2.value = ordena[i]
                cell2.font = ft
                cell2.border = thin_border

                total += int(ordena[i])
                linha += 1
                break
        del dic_Civil[names]
    ws.cell(row=linha, column=1).value = "Total"
    ws.cell(row=linha, column=1).font = ft
    ws.cell(row=linha, column=1).border = thin_border

    ws.cell(row=linha, column=2).value = total
    ws.cell(row=linha, column=2).font = ft
    ws.cell(row=linha, column=2).border = thin_border

    ########################################################################################################################

    ws.cell(row=6, column=4).value = "Técnicos Elétrica"
    ws.cell(row=6, column=4).border = thin_border
    ws.cell(row=6, column=4).font = ft

    ws.cell(row=6, column=5).value = "Solicitações"
    ws.cell(row=6, column=5).border = thin_border
    ws.cell(row=6, column=5).font = ft

    total = 0
    linha = 7

    ordena = list(reversed(sorted(dic_Eletrica.values())))
    for i in range(len(dic_Eletrica.keys())):
        for names in dic_Eletrica.keys():
            if dic_Eletrica[names] == ordena[i]:
                cell = ws.cell(row=linha, column=4)
                cell.value = names
                cell.font = ft
                cell.border = thin_border

                cell2 = ws.cell(row=linha, column=5)
                cell2.value = ordena[i]
                cell2.font = ft
                cell2.border = thin_border

                total += int(ordena[i])
                linha += 1
                break
        del dic_Eletrica[names]

    ws.cell(row=linha, column=4).value = "Total"
    ws.cell(row=linha, column=4).font = ft
    ws.cell(row=linha, column=4).border = thin_border

    ws.cell(row=linha, column=5).value = total
    ws.cell(row=linha, column=5).font = ft
    ws.cell(row=linha, column=5).border = thin_border

    ########################################################################################################################

    ws.cell(row=6, column=7).value = "Técnicos Mecânica"
    ws.cell(row=6, column=7).border = thin_border
    ws.cell(row=6, column=7).font = ft

    ws.cell(row=6, column=8).value = "Solicitações"
    ws.cell(row=6, column=8).border = thin_border
    ws.cell(row=6, column=8).font = ft

    total = 0
    linha = 7

    ordena = list(reversed(sorted(dic_Mecanica.values())))
    for i in range(len(dic_Mecanica.keys())):
        for names in dic_Mecanica.keys():
            if dic_Mecanica[names] == ordena[i]:
                cell = ws.cell(row=linha, column=7)
                cell.value = names
                cell.font = ft
                cell.border = thin_border

                cell2 = ws.cell(row=linha, column=8)
                cell2.value = ordena[i]
                cell2.font = ft
                cell2.border = thin_border

                total += int(ordena[i])
                linha += 1
                break
        del dic_Mecanica[names]

    ws.cell(row=linha, column=7).value = "Total"
    ws.cell(row=linha, column=7).font = ft
    ws.cell(row=linha, column=7).border = thin_border

    ws.cell(row=linha, column=8).value = total
    ws.cell(row=linha, column=8).font = ft
    ws.cell(row=linha, column=8).border = thin_border

    """""""""""""""""""""""""""
    Chamados Abertos
    """""
    redFill = PatternFill(start_color='e2fe13', end_color='e2c813',
                          fill_type='solid')
    ws_new = wb.create_sheet(title="Chamados Abertos")

    ws_new.merge_cells('A1:H1')
    cell = ws_new.cell(row=1, column=1)
    cell.value = 'Chamados Abertos'
    cell.fill = redFill

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for i in range(1, 9, 1):
        ws_new.cell(row=1, column=i).border = thin_border
        ws_new.cell(row=2, column=i).border = thin_border
    lst2 = ["Solicitações ", "Situação", "Localização", "Técnico", "Loja", "Inicio", "Descrição",
            "Item"]
    z = 1
    ft = Font(color=colors.WHITE)

    blackFill = PatternFill(start_color='010204',
                            end_color='010204',
                            fill_type='solid')

    for name in lst2:
        ws_new.cell(row=2, column=z).value = name
        ws_new.cell(row=2, column=z).font = ft
        ws_new.cell(row=2, column=z).fill = blackFill
        z += 1

    cell.alignment = Alignment(horizontal='center', vertical='center')








    wb.save('Relatorio/Relatorio_InternoVsExterno.xlsx')
    wb.close()


def gerarColunas():
    import string
    a = string.ascii_uppercase
    lst = []
    for i in a:
        lst.append(i)
    lst2 = []
    for j in range(5):
        for i in a:
            lst2.append(lst[j] + i)
    lst += lst2
    return lst


def calc(hour, hours2):
    date_format = "%m-%d-%Y %H:%M:%S"
    time1 = datetime.strptime('8-01-2008 ' + hour, date_format)
    time2 = datetime.strptime('8-01-2008 ' + hours2, date_format)
    diff = time2 - time1
    minutes = diff.seconds / 60
    print(str(minutes) + ' Minutes')


NAO = 0
SIM = 0
TOTAL = 0
CANCELADAS = 0
ABERTOS_HJ = 0
APROVAR = 0
hour = str(datetime.now().hour) + ":" + str(datetime.now().minute) + ":" + str(datetime.now().second)

wb = load_workbook(filename='Resultado.xlsx', read_only=True)
ws = wb['Resultado da consulta de solici']

colunas = gerarColunas()

for i in colunas:
    coluna = ws[i + "1"]
    if coluna.value == "Solicitação":
        coluna_solicitacao = i

    elif coluna.value == "Situação":
        coluna_situacao = i

    elif coluna.value == "Localização":
        coluna_localizacao = i

    elif coluna.value == "Início":
        coluna_inicio = i

    elif coluna.value == "Fim":
        coluna_fim = i

    elif coluna.value == "SLACONSUMIDOEXECUCAOTEC - solicitacao_infraestrutura":
        coluna_slaexe = i

    elif coluna.value == "SLACONSUMIDOACOMPANHAMENTO - solicitacao_infraestrutura":
        coluna_slaacom = i

    elif coluna.value == "DESCRICAOITEM - solicitacao_infraestrutura":
        coluna_item = i

    elif coluna.value == "TOTALHOURSSLA - solicitacao_infraestrutura":
        coluna_sla = i

    elif coluna.value == "CODIGOTECNICO - solicitacao_infraestrutura":
        coluna_tec = i

    elif coluna.value == "Atividade - solicitacao_infraestrutura - Aprovar conclusão - Conclusão":
        coluna_AprovarConclusao = i

    elif coluna.value == "DATAENTRADACLASSIFICACAO - solicitacao_infraestrutura":
        coluna_Classificacao = i

    elif coluna.value == "Atividade - solicitacao_infraestrutura - Início - Conclusão":
        if "E" not in i:
            coluna_inicioConclusao = i

    elif coluna.value == "DATASAIDAEXECUCAOTEC - solicitacao_infraestrutura":
        coluna_SaidaExecTec = i

    elif coluna.value == "DATASAIDAACOMPANHAMENTO - solicitacao_infraestrutura":
        coluna_SaidaAcompanhamento = i

    elif coluna.value == "Atividade - solicitacao_infraestrutura - Execução 2º técnico? - Conclusão":
        coluna_Exec2tec = i

    elif coluna.value == "DESCRICAO - solicitacao_infraestrutura":
        coluna_Descri = i

    elif coluna.value == "RADIOAPROVREJ - solicitacao_infraestrutura":
        coluna_Rejeitado = i

    elif coluna.value == "Atividade - solicitacao_infraestrutura - Aprovação? - Conclusão":
        coluna_Aprovacao = i


def main():
    arq = open("Tecnicos.txt", 'rt')
    lines = arq.readlines()
    for line in lines:
        dic.update({line.split(";")[0]: line.split(";")[1].strip()})
    arq.close()

    from openpyxl import Workbook
    wb_new = Workbook()
    ws_new = wb_new['Sheet']
    ws_new.title = "Dados"

    ft = Font(color=colors.WHITE)
    blackFill = PatternFill(start_color='010204',
                            end_color='010204',
                            fill_type='solid')

    lista_names = ['Solicitações', 'Situação', 'Localização', 'Técnico', 'Inicio', 'Fim', 'Item', 'SLA',
                   'SLA Consumido', "Dentro do prazo?", "Inicio contador SLA", "Fim contador SLA"]
    z = 1
    for name in lista_names:
        ws_new.cell(row=1, column=z).value = name
        ws_new.cell(row=1, column=z).font = ft
        ws_new.cell(row=1, column=z).fill = blackFill
        z += 1

    wb_new.save("Relatorio/Relatorio_InternoVsExterno.xlsx")
    wb_new.close()

    import threading
    percorre = 20
    for i in range(1, 4, 1):
        T1 = threading.Thread(target=GeraExcel, args=(percorre * i, i, percorre))
        T1.start()


Cont = 0


def GeraExcel(percorre, thread, percorre2):
    wb = load_workbook(filename='Resultado.xlsx', read_only=True, data_only=True)
    ws = wb['Resultado da consulta de solici']
    Lista_Salvar = []
    if thread == 1:
        j = percorre - percorre2 + 2
    else:
        j = percorre - percorre2
    for i in range(j, percorre, 1):

        global CANCELADAS, APROVAR, ABERTOS_HJ, NAO, SIM, TOTAL
        solicitacoes = str(int(ws[coluna_solicitacao + str(i)].value)).strip()

        if solicitacoes != "":
            print(thread, solicitacoes)
            localizacao = str(ws[coluna_localizacao + str(i)].value).strip()
            if localizacao == "Classificação" or localizacao == "Acompanhamento da Execução" or localizacao == "Acompanharmento da Execução" or localizacao == "Execução da Solicitação" or localizacao == "Analise?":
                situacao = "Em aberto"
            elif localizacao == "Aprovar":
                situacao = "Aprovação"
                APROVAR += 1
            elif localizacao == "Cancelada":
                situacao = "Cancelada"
                CANCELADAS += 1
            else:
                situacao = "Fechado"

            try:
                tecnico = dic[ws[coluna_tec + str(i)].value]
            except Exception:
                tecnico = ws[coluna_tec + str(i)].value

            item = ws[coluna_item + str(i)].value
            SLA = ws[coluna_sla + str(i)].value
            inicio = ws[coluna_inicio + str(i)].value
            if inicio.split("/")[1] == "10":
                ABERTOS_HJ += 1

            if situacao == "Fechado":
                rejei = ""
                DataEntrada_classificacao = str(
                    ws[coluna_Classificacao + str(
                        i)].value).strip()  # DATAENTRADACLASSIFICACAO - solicitacao_infraestrutura

                Data_Entrada_AprovarConclusao = PadronizaDate2(str(ws[coluna_AprovarConclusao + str(
                    i)].value).strip())  # Atividade - solicitacao_infraestrutura - Aprovar conclusão - Conclusão

                Item_Sla = str(ws[coluna_sla + str(i)].value).strip().split(":")[
                    0]  # TOTALHOURSSLA - solicitacao_infraestrutura

                if DataEntrada_classificacao != "" and DataEntrada_classificacao is not None and DataEntrada_classificacao != " " and DataEntrada_classificacao != "None":

                    dataEntrada = PadronizaDate(DataEntrada_classificacao)

                    if Data_Entrada_AprovarConclusao != "" and Data_Entrada_AprovarConclusao is not None and Data_Entrada_AprovarConclusao != "None":
                        dif = calcHours(dataEntrada + ":00", Data_Entrada_AprovarConclusao)

                    else:
                        try:
                            Data_Entrada_AprovarConclusao = PadronizaDate(ws[coluna_SaidaExecTec + str(i)].value)
                            try:
                                if Data_Entrada_AprovarConclusao < PadronizaDate(str(ws[coluna_SaidaAcompanhamento + str(i)].value).strip()):
                                    Data_Entrada_AprovarConclusao = PadronizaDate(
                                        str(ws[coluna_SaidaAcompanhamento + str(i)].value).strip())
                            except Exception:
                                pass

                        except Exception:
                            try:
                                Data_Entrada_AprovarConclusao = PadronizaDate(ws[coluna_SaidaAcompanhamento + str(i)].value)
                            except Exception:
                                rejei = "rejeitada"

                        dif = round(calcHours(dataEntrada + ":00", Data_Entrada_AprovarConclusao + ":00"), 2)
                        Data_Entrada_AprovarConclusao += ":00"
                        if dif <= 0:
                            try:
                                Data_Entrada_AprovarConclusao = str(ws[coluna_Exec2tec + str(i)].value).strip()
                                print(dataEntrada,Data_Entrada_AprovarConclusao)
                                dif = round(calcHours(dataEntrada + ":00", PadronizaDate2(Data_Entrada_AprovarConclusao)), 2)
                            except Exception:
                                rejei = str(ws[coluna_Rejeitado + str(i)].value).strip()
                    if rejei != "rejeitada":
                        if dif >= int(Item_Sla):
                            dif = str(dif).split(".")[0] + ":" + str(round(float("0." + str(dif).split(".")[1]) * 60))
                            Lista_Salvar.append([solicitacoes, situacao, localizacao, tecnico, inicio, item, SLA, dif,
                                                 "Não", dataEntrada + ":00", Data_Entrada_AprovarConclusao])
                            NAO += 1
                        else:
                            dif = str(dif).split(".")[0] + ":" + str(round(float("0." + str(dif).split(".")[1]) * 60))
                            Lista_Salvar.append(
                                [solicitacoes, situacao, localizacao, tecnico, inicio, item, SLA, dif,
                                 "Sim", dataEntrada + ":00", Data_Entrada_AprovarConclusao])
                            SIM += 1
                    else:
                        Lista_Salvar.append(
                            [solicitacoes, situacao, localizacao, tecnico, inicio, item, SLA, "-",
                             "-", "-", "-"])

                else:
                    """ Com orgão fiscalizador ativo """
                    DataSaida_Inicio = str(
                        ws[coluna_inicioConclusao + str(
                            i)].value).strip()  # Atividade - solicitacao_infraestrutura - Início - Conclusão

                    DataSaida_Exec = str(
                        ws[coluna_SaidaExecTec + str(i)].value).strip()  # DATASAIDAEXECUCAOTEC - solicitacao_infraestrutura

                    if DataSaida_Exec != "" and DataSaida_Exec != None and DataSaida_Exec != " " and DataSaida_Exec != "None":

                        dataSaida_Exec = PadronizaDate(DataSaida_Exec) + ":00"
                        try:
                            if dataSaida_Exec < PadronizaDate(str(ws[coluna_SaidaAcompanhamento + str(i)].value).strip()):
                                dataSaida_Exec = PadronizaDate(str(ws[coluna_SaidaAcompanhamento + str(i)].value).strip())
                        except Exception:
                            pass

                    else:
                        try:
                            DataSaida_Exec = str(
                                ws[coluna_SaidaAcompanhamento + str(
                                    i)].value).strip()  # DATASAIDAACOMPANHAMENTO - solicitacao_infraestrutura
                            dataSaida_Exec = PadronizaDate(DataSaida_Exec) + ":00"
                        except Exception:
                            rejei = str(ws[coluna_Rejeitado + str(i)].value).strip()

                        try:
                            if dataSaida_Exec < PadronizaDate(str(ws[coluna_SaidaExecTec + str(i)].value).strip()):
                                dataSaida_Exec = PadronizaDate(str(ws[coluna_SaidaExecTec + str(i)].value).strip())
                        except Exception:
                            pass

                    if rejei != "rejeitada":
                        dataSaida_Inicio = PadronizaDate2(DataSaida_Inicio)
                        if dataSaida_Inicio != "None" and dataSaida_Inicio != None and dataSaida_Inicio != "":
                            dif = round(calcHours(dataSaida_Inicio, dataSaida_Exec), 2)
                        else:
                            dataSaida_Inicio = str(ws[coluna_Aprovacao + str(i)].value).strip()
                            dif = round(calcHours(PadronizaDate2(dataSaida_Inicio), dataSaida_Exec), 2)


                        if dif >= int(Item_Sla):
                            dif = str(dif).split(".")[0] + ":" + str(round(float("0." + str(dif).split(".")[1]) * 60))
                            Lista_Salvar.append(
                                [solicitacoes, situacao, localizacao, tecnico, inicio, item, SLA, dif,
                                 "Não", dataSaida_Inicio, dataSaida_Exec])
                            NAO += 1
                        else:
                            dif = str(dif).split(".")[0] + ":" + str(round(float("0." + str(dif).split(".")[1]) * 60))
                            Lista_Salvar.append(
                                [solicitacoes, situacao, localizacao, tecnico, inicio, item, SLA, dif,
                                 "Sim", dataSaida_Inicio, dataSaida_Exec])
                            SIM += 1
                    else:
                        Lista_Salvar.append(
                            [solicitacoes, situacao, localizacao, tecnico, inicio, item, SLA, "-",
                             "-", "-", "-"])

            else:
                Lista_Salvar.append(
                    [solicitacoes, situacao, localizacao, tecnico, inicio, item, SLA, "-",
                     "-", "-", "-"])

            TOTAL += 1

    print(str(thread) + " thread finalizada")

    Finaliza(thread, Lista_Salvar)


lst = []
Lista_Salvar2 = []


def Finaliza(thread, Lista_Salvar):
    global Lista_Salvar2
    lst.append(thread)

    Lista_Salvar2 += Lista_Salvar

    if len(lst) == 3:
        wb_new = load_workbook(filename='Relatorio/Relatorio_InternoVsExterno.xlsx', read_only=False)
        ws_new = wb_new["Dados"]
        global NAO, SIM, TOTAL

        for list in Lista_Salvar2:
            ws_new.append(list)
        ws_new.auto_filter.ref = 'A1:J9'

        dims = {}
        for row in ws_new.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws_new.column_dimensions[col].width = value

        redFill = PatternFill(start_color='e2fe13', end_color='e2c813',
                              fill_type='solid')

        ws_new = wb_new.create_sheet(title="Informações")
        ws_new.merge_cells('A1:H1')
        cell = ws_new.cell(row=1, column=1)
        cell.value = 'Geral'
        cell.fill = redFill

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for i in range(1, 9, 1):
            ws_new.cell(row=1, column=i).border = thin_border
            ws_new.cell(row=2, column=i).border = thin_border
        lst2 = ["Total ", "Total Ativos", "Abertos em Dezembro", "Total de solicitações finalizadas",
                "Percentual de Atendimento", "Finalizadas dentro do prazo", "Finalizadas fora do prazo",
                "Canceladas"]
        z = 1
        ft = Font(color=colors.WHITE)

        blackFill = PatternFill(start_color='010204',
                                end_color='010204',
                                fill_type='solid')

        for name in lst2:
            ws_new.cell(row=2, column=z).value = name
            ws_new.cell(row=2, column=z).font = ft
            ws_new.cell(row=2, column=z).fill = blackFill
            z += 1

        cell.alignment = Alignment(horizontal='center', vertical='center')

        ws_new.append(
            [TOTAL, TOTAL - CANCELADAS, ABERTOS_HJ, SIM + NAO,
             str(round(((SIM + NAO) * 100) / TOTAL - CANCELADAS)) + " %", SIM,
             NAO, CANCELADAS])

        dims = {}
        for row in ws_new.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws_new.column_dimensions[col].width = value

        wb_new.save('Relatorio/Relatorio_InternoVsExterno.xlsx')
        wb_new.close()
        hour2 = str(datetime.now().hour) + ":" + str(datetime.now().minute) + ":" + str(datetime.now().second)
        Juncao()

        calc(hour, hour2)

main()
